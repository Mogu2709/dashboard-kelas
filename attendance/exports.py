import io
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Count

from .models import Pertemuan, Attendance


def get_rekap_data():
    """Helper: ambil data rekap mahasiswa."""
    total_pertemuan = Pertemuan.objects.count()
    mahasiswa = User.objects.filter(is_superuser=False).annotate(
        total_hadir=Count('attendance')
    ).order_by('username')

    data = []
    for mhs in mahasiswa:
        persen = round((mhs.total_hadir / total_pertemuan) * 100, 1) if total_pertemuan > 0 else 0
        if persen >= 75:
            status = 'Aman'
        elif persen >= 50:
            status = 'Perhatian'
        else:
            status = 'Bahaya'
        data.append({
            'username': mhs.username,
            'total_hadir': mhs.total_hadir,
            'total_pertemuan': total_pertemuan,
            'persentase': persen,
            'status': status,
        })
    return data, total_pertemuan


# ─── EXPORT EXCEL ────────────────────────────────────────────────────────────

@login_required
def export_excel(request):
    if not request.user.is_superuser:
        from django.shortcuts import redirect
        return redirect('pertemuan_list')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data, total_pertemuan = get_rekap_data()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Kehadiran"

    # ── Styles ──
    header_font   = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill   = PatternFill('solid', fgColor='1a1a1a')
    center        = Alignment(horizontal='center', vertical='center')
    left          = Alignment(horizontal='left',   vertical='center')

    thin = Side(style='thin', color='E9E9E7')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_aman      = PatternFill('solid', fgColor='EEFAF4')
    fill_perhatian = PatternFill('solid', fgColor='FEFCE8')
    fill_bahaya    = PatternFill('solid', fgColor='FEF2F2')

    font_aman      = Font(name='Calibri', color='1A7F4B', size=11)
    font_perhatian = Font(name='Calibri', color='854D0E', size=11)
    font_bahaya    = Font(name='Calibri', color='B91C1C', size=11)

    # ── Title ──
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = 'Rekap Kehadiran Mahasiswa'
    title_cell.font = Font(name='Calibri', bold=True, size=14)
    title_cell.alignment = center
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:E2')
    sub_cell = ws['A2']
    sub_cell.value = f'Total Pertemuan: {total_pertemuan}'
    sub_cell.font = Font(name='Calibri', color='9B9B9B', size=10)
    sub_cell.alignment = center
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 8  # spacer

    # ── Header row ──
    headers = ['No', 'Username', 'Total Hadir', 'Persentase', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[4].height = 22

    # ── Data rows ──
    for i, row in enumerate(data, 1):
        r = i + 4
        ws.row_dimensions[r].height = 20

        status = row['status']
        s_fill = fill_aman if status == 'Aman' else (fill_perhatian if status == 'Perhatian' else fill_bahaya)
        s_font = font_aman if status == 'Aman' else (font_perhatian if status == 'Perhatian' else font_bahaya)

        values = [i, row['username'], f"{row['total_hadir']}/{total_pertemuan}", f"{row['persentase']}%", status]
        aligns = [center, left, center, center, center]

        for col, (val, aln) in enumerate(zip(values, aligns), 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = aln
            cell.border = border
            cell.font = Font(name='Calibri', size=11)
            if col == 5:
                cell.fill = s_fill
                cell.font = s_font

    # ── Column widths ──
    for col, width in zip(range(1, 6), [6, 28, 16, 16, 14]):
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Response ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="rekap_kehadiran.xlsx"'
    return response


# ─── EXPORT PDF ──────────────────────────────────────────────────────────────

@login_required
def export_pdf(request):
    if not request.user.is_superuser:
        from django.shortcuts import redirect
        return redirect('pertemuan_list')

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    data, total_pertemuan = get_rekap_data()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    elements = []

    # ── Title ──
    title_style = ParagraphStyle('title',
        fontName='Helvetica-Bold', fontSize=16,
        textColor=colors.HexColor('#1a1a1a'),
        alignment=TA_CENTER, spaceAfter=4)
    sub_style = ParagraphStyle('sub',
        fontName='Helvetica', fontSize=10,
        textColor=colors.HexColor('#9b9b9b'),
        alignment=TA_CENTER, spaceAfter=20)

    elements.append(Paragraph('Rekap Kehadiran Mahasiswa', title_style))
    elements.append(Paragraph(f'Total Pertemuan: {total_pertemuan}', sub_style))

    # ── Table data ──
    col_widths = [1.2*cm, 6.5*cm, 3*cm, 3*cm, 3*cm]

    table_data = [['No', 'Username', 'Total Hadir', 'Persentase', 'Status']]
    for i, row in enumerate(data, 1):
        table_data.append([
            str(i),
            row['username'],
            f"{row['total_hadir']}/{total_pertemuan}",
            f"{row['persentase']}%",
            row['status'],
        ])

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Base style
    ts = TableStyle([
        # Header
        ('BACKGROUND',  (0,0), (-1,0), colors.HexColor('#1a1a1a')),
        ('TEXTCOLOR',   (0,0), (-1,0), colors.white),
        ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,0), 9),
        ('ALIGN',       (0,0), (-1,0), 'CENTER'),
        ('TOPPADDING',  (0,0), (-1,0), 8),
        ('BOTTOMPADDING',(0,0),(-1,0), 8),

        # Body
        ('FONTNAME',    (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE',    (0,1), (-1,-1), 9),
        ('ALIGN',       (0,1), (0,-1), 'CENTER'),   # No
        ('ALIGN',       (1,1), (1,-1), 'LEFT'),      # Username
        ('ALIGN',       (2,1), (-1,-1), 'CENTER'),   # rest
        ('TOPPADDING',  (0,1), (-1,-1), 7),
        ('BOTTOMPADDING',(0,1),(-1,-1), 7),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f7f7f5')]),

        # Grid
        ('GRID',        (0,0), (-1,-1), 0.5, colors.HexColor('#e9e9e7')),
        ('LINEBELOW',   (0,0), (-1,0), 1, colors.HexColor('#1a1a1a')),
    ])

    # Status colors per row
    for i, row in enumerate(data, 1):
        if row['status'] == 'Aman':
            ts.add('TEXTCOLOR', (4, i), (4, i), colors.HexColor('#1a7f4b'))
            ts.add('BACKGROUND', (4, i), (4, i), colors.HexColor('#eefaf4'))
        elif row['status'] == 'Perhatian':
            ts.add('TEXTCOLOR', (4, i), (4, i), colors.HexColor('#854d0e'))
            ts.add('BACKGROUND', (4, i), (4, i), colors.HexColor('#fefce8'))
        else:
            ts.add('TEXTCOLOR', (4, i), (4, i), colors.HexColor('#b91c1c'))
            ts.add('BACKGROUND', (4, i), (4, i), colors.HexColor('#fef2f2'))

    table.setStyle(ts)
    elements.append(table)

    # ── Footer ──
    from django.utils import timezone
    now_str = timezone.now().strftime('%d %B %Y, %H:%M')
    footer_style = ParagraphStyle('footer',
        fontName='Helvetica', fontSize=8,
        textColor=colors.HexColor('#c4c4c0'),
        alignment=TA_LEFT, spaceBefore=16)
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph(f'Dicetak pada: {now_str}', footer_style))

    doc.build(elements)
    buf.seek(0)

    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="rekap_kehadiran.pdf"'
    return response
