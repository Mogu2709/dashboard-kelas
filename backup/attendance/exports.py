import io
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Count, Q

from .models import Pertemuan, Attendance, IzinAbsen


def get_rekap_data():
    """
    Helper: ambil data rekap mahasiswa.
    FIX 1: Filter hanya user approved (bukan pending/rejected).
    FIX 2: Hitung izin & sakit agar konsisten dengan views.rekap_mahasiswa.
    FITUR: Sertakan nama_lengkap dan NIM dari profile.
    """
    total_pertemuan = Pertemuan.objects.count()
    pertemuan_ids = list(Pertemuan.objects.values_list('id', flat=True))

    # FIX 1: Filter hanya mahasiswa yang sudah approved
    mahasiswa = (
        User.objects
        .filter(is_superuser=False, profile__status='approved')
        .select_related('profile')
        .annotate(total_hadir=Count(
            'attendance',
            filter=Q(attendance__pertemuan_id__in=pertemuan_ids)
        ))
        .order_by('username')
    )

    # FIX 2: Izin & sakit per user (approved only) — satu query
    izin_map = {}
    if pertemuan_ids:
        for row in (
            IzinAbsen.objects
            .filter(pertemuan_id__in=pertemuan_ids, status='approved')
            .values('user_id', 'jenis')
            .annotate(jumlah=Count('id'))
        ):
            uid = row['user_id']
            if uid not in izin_map:
                izin_map[uid] = {'izin': 0, 'sakit': 0}
            if row['jenis'] in ('izin', 'sakit'):
                izin_map[uid][row['jenis']] = row['jumlah']

    data = []
    for mhs in mahasiswa:
        tot_hadir = mhs.total_hadir
        tot_izin  = izin_map.get(mhs.id, {}).get('izin', 0)
        tot_sakit = izin_map.get(mhs.id, {}).get('sakit', 0)
        efektif   = tot_hadir + tot_izin + tot_sakit
        persen    = round((efektif / total_pertemuan) * 100, 1) if total_pertemuan > 0 else 0
        persen_murni = round((tot_hadir / total_pertemuan) * 100, 1) if total_pertemuan > 0 else 0

        if persen >= 75:
            status = 'Aman'
        elif persen >= 50:
            status = 'Perhatian'
        else:
            status = 'Bahaya'

        try:
            nama = mhs.profile.nama_lengkap or mhs.username
            nim  = mhs.profile.nim or '-'
        except Exception:
            nama = mhs.username
            nim  = '-'

        data.append({
            'username': mhs.username,
            'nama_lengkap': nama,
            'nim': nim,
            'total_hadir': tot_hadir,
            'total_izin': tot_izin,
            'total_sakit': tot_sakit,
            'total_pertemuan': total_pertemuan,
            'persentase': persen,
            'persentase_murni': persen_murni,
            'status': status,
        })
    return data, total_pertemuan


# ─── EXPORT EXCEL ─────────────────────────────────────────────────────────────

@login_required
def export_excel(request):
    if not request.user.is_superuser:
        from django.shortcuts import redirect
        return redirect('pertemuan_list')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.utils import timezone

    data, total_pertemuan = get_rekap_data()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Kehadiran"

    header_font    = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill    = PatternFill('solid', fgColor='1a1a1a')
    center         = Alignment(horizontal='center', vertical='center')
    left           = Alignment(horizontal='left',   vertical='center')
    thin   = Side(style='thin', color='E9E9E7')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_aman      = PatternFill('solid', fgColor='EEFAF4')
    fill_perhatian = PatternFill('solid', fgColor='FEFCE8')
    fill_bahaya    = PatternFill('solid', fgColor='FEF2F2')
    font_aman      = Font(name='Calibri', color='1A7F4B', size=11)
    font_perhatian = Font(name='Calibri', color='854D0E', size=11)
    font_bahaya    = Font(name='Calibri', color='B91C1C', size=11)

    num_cols = 9
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    ws['A1'].value     = 'Rekap Kehadiran Mahasiswa'
    ws['A1'].font      = Font(name='Calibri', bold=True, size=14)
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    ws['A2'].value     = f'Total Pertemuan: {total_pertemuan}  |  Dicetak: {timezone.now().strftime("%d %B %Y %H:%M")}'
    ws['A2'].font      = Font(name='Calibri', color='9B9B9B', size=10)
    ws['A2'].alignment = center
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8

    headers = ['No', 'NIM', 'Nama Lengkap', 'Username', 'Hadir', 'Izin', 'Sakit', 'Efektif (%)', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = border
    ws.row_dimensions[4].height = 22

    for i, row in enumerate(data, 1):
        r = i + 4
        ws.row_dimensions[r].height = 20
        status = row['status']
        s_fill = fill_aman if status == 'Aman' else (fill_perhatian if status == 'Perhatian' else fill_bahaya)
        s_font = font_aman if status == 'Aman' else (font_perhatian if status == 'Perhatian' else font_bahaya)
        values = [
            i, row['nim'], row['nama_lengkap'], row['username'],
            f"{row['total_hadir']}/{total_pertemuan}",
            row['total_izin'], row['total_sakit'],
            f"{row['persentase']}%", status,
        ]
        aligns = [center, left, left, left, center, center, center, center, center]
        for col, (val, aln) in enumerate(zip(values, aligns), 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = aln; cell.border = border
            cell.font = Font(name='Calibri', size=11)
            if col == 9:
                cell.fill = s_fill; cell.font = s_font

    for col, width in enumerate([5, 16, 28, 18, 14, 8, 8, 14, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="rekap_kehadiran.xlsx"'
    return response


# ─── EXPORT PDF ───────────────────────────────────────────────────────────────

@login_required
def export_pdf(request):
    if not request.user.is_superuser:
        from django.shortcuts import redirect
        return redirect('pertemuan_list')

    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from django.utils import timezone

    data, total_pertemuan = get_rekap_data()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )
    elements = []

    title_style = ParagraphStyle('title', fontName='Helvetica-Bold', fontSize=16,
        textColor=colors.HexColor('#1a1a1a'), alignment=TA_CENTER, spaceAfter=4)
    sub_style = ParagraphStyle('sub', fontName='Helvetica', fontSize=9,
        textColor=colors.HexColor('#9b9b9b'), alignment=TA_CENTER, spaceAfter=20)

    elements.append(Paragraph('Rekap Kehadiran Mahasiswa', title_style))
    elements.append(Paragraph(
        f'Total Pertemuan: {total_pertemuan}  |  Dicetak: {timezone.now().strftime("%d %B %Y %H:%M")}',
        sub_style
    ))

    col_widths = [0.8*cm, 3*cm, 6*cm, 3*cm, 2.5*cm, 1.5*cm, 1.5*cm, 2.8*cm, 2.5*cm]
    table_data = [['No', 'NIM', 'Nama Lengkap', 'Username', 'Hadir', 'Izin', 'Sakit', 'Efektif (%)', 'Status']]
    for i, row in enumerate(data, 1):
        table_data.append([
            str(i), row['nim'], row['nama_lengkap'], row['username'],
            f"{row['total_hadir']}/{total_pertemuan}",
            str(row['total_izin']), str(row['total_sakit']),
            f"{row['persentase']}%", row['status'],
        ])

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    ts = TableStyle([
        ('BACKGROUND',   (0, 0), (-1, 0), colors.HexColor('#1a1a1a')),
        ('TEXTCOLOR',    (0, 0), (-1, 0), colors.white),
        ('FONTNAME',     (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0), (-1, 0), 9),
        ('ALIGN',        (0, 0), (-1, 0), 'CENTER'),
        ('TOPPADDING',   (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING',(0, 0), (-1, 0), 8),
        ('FONTNAME',     (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',     (0, 1), (-1, -1), 8),
        ('ALIGN',        (0, 1), (0, -1), 'CENTER'),
        ('ALIGN',        (1, 1), (3, -1), 'LEFT'),
        ('ALIGN',        (4, 1), (-1, -1), 'CENTER'),
        ('TOPPADDING',   (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING',(0, 1), (-1, -1), 6),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#f7f7f5')]),
        ('GRID',         (0, 0), (-1, -1), 0.5, colors.HexColor('#e9e9e7')),
        ('LINEBELOW',    (0, 0), (-1, 0), 1, colors.HexColor('#1a1a1a')),
    ])
    for i, row in enumerate(data, 1):
        if row['status'] == 'Aman':
            ts.add('TEXTCOLOR',  (8, i), (8, i), colors.HexColor('#1a7f4b'))
            ts.add('BACKGROUND', (8, i), (8, i), colors.HexColor('#eefaf4'))
        elif row['status'] == 'Perhatian':
            ts.add('TEXTCOLOR',  (8, i), (8, i), colors.HexColor('#854d0e'))
            ts.add('BACKGROUND', (8, i), (8, i), colors.HexColor('#fefce8'))
        else:
            ts.add('TEXTCOLOR',  (8, i), (8, i), colors.HexColor('#b91c1c'))
            ts.add('BACKGROUND', (8, i), (8, i), colors.HexColor('#fef2f2'))
    table.setStyle(ts)
    elements.append(table)

    footer_style = ParagraphStyle('footer', fontName='Helvetica', fontSize=8,
        textColor=colors.HexColor('#c4c4c0'), alignment=TA_LEFT, spaceBefore=16)
    elements.append(Spacer(1, 0.3*cm))
    elements.append(Paragraph('Dashboard Kelas — Data hanya mencakup mahasiswa dengan status disetujui.', footer_style))

    doc.build(elements)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="rekap_kehadiran.pdf"'
    return response
